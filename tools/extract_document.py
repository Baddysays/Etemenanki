#!/usr/bin/env python3
"""Extract plain text from txt/md/docx/pdf for Etemenanki."""
from __future__ import annotations

import json
import html
import re
import sys
from pathlib import Path


def decode_text_bytes(data: bytes) -> tuple[str, str]:
    if data.startswith(b"\xef\xbb\xbf"):
        return data[3:].decode("utf-8", errors="replace"), "utf-8-sig"
    if data.startswith(b"\xff\xfe") and len(data) >= 2:
        return data[2:].decode("utf-16-le", errors="replace"), "utf-16-le"
    if data.startswith(b"\xfe\xff") and len(data) >= 2:
        return data[2:].decode("utf-16-be", errors="replace"), "utf-16-be"
    try:
        return data.decode("utf-8"), "utf-8"
    except UnicodeDecodeError:
        pass
    for enc in ("cp1251", "koi8-r", "latin-1"):
        try:
            return data.decode(enc), enc
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace"), "utf-8-replace"


def extract_txt(path: Path) -> tuple[list[str], str]:
    data = path.read_bytes()
    text, enc = decode_text_bytes(data)
    pages = [text] if text.strip() else [""]
    return pages, enc


def extract_docx(path: Path) -> tuple[list[str], str, dict]:
    from docx import Document

    doc = Document(str(path))
    paragraphs = [p.text for p in doc.paragraphs]
    segments = []
    for idx, text in enumerate(paragraphs):
        cleaned = text.strip()
        if not cleaned:
            continue
        segments.append({"id": f"p{idx}", "index": idx, "text": cleaned})
    full = "\n".join(paragraphs).strip()
    page_count = max(1, min(120, len(paragraphs) // 8 + 1))
    chunk = max(1, len(paragraphs) // page_count)
    pages: list[str] = []
    for i in range(0, len(paragraphs), chunk):
        pages.append("\n".join(paragraphs[i : i + chunk]).strip())
    if not pages:
        pages = [full]
    meta = {"segments": segments, "format": "docx"}
    return pages, "docx", meta


def _chunk_segments(segments: list[dict], per_page: int = 40) -> list[str]:
    pages: list[str] = []
    for i in range(0, len(segments), per_page):
        block = segments[i : i + per_page]
        lines = [f"{item['id']}|{item['text']}" for item in block if item.get("text")]
        if lines:
            pages.append("\n".join(lines))
    return pages or [""]


def extract_subtitles(path: Path) -> tuple[list[str], str, dict]:
    raw, enc = decode_text_bytes(path.read_bytes())
    suffix = path.suffix.lower()
    segments: list[dict] = []

    if suffix == ".vtt":
        lines = raw.splitlines()
        idx = 0
        i = 0
        while i < len(lines):
            line = lines[i].strip()
            if "-->" in line:
                timing = line.replace(".", ",")
                text_lines: list[str] = []
                i += 1
                while i < len(lines) and lines[i].strip():
                    text_lines.append(lines[i].strip())
                    i += 1
                text = "\n".join(text_lines).strip()
                if text:
                    segments.append(
                        {
                            "id": f"c{idx}",
                            "index": idx,
                            "start": timing.split("-->")[0].strip(),
                            "end": timing.split("-->")[1].strip() if "-->" in timing else "",
                            "text": text,
                        }
                    )
                    idx += 1
            i += 1
    elif suffix == ".ass":
        for line in raw.splitlines():
            if not line.startswith("Dialogue:"):
                continue
            parts = line.split(",", 9)
            if len(parts) < 10:
                continue
            text = parts[9].strip()
            if not text:
                continue
            segments.append(
                {
                    "id": f"c{len(segments)}",
                    "index": len(segments),
                    "start": parts[1].strip(),
                    "end": parts[2].strip(),
                    "text": text,
                    "ass": True,
                }
            )
    else:
        blocks = re.split(r"\n\s*\n", raw.strip())
        for block in blocks:
            lines = [ln.strip() for ln in block.splitlines() if ln.strip()]
            if len(lines) < 3:
                continue
            if not re.match(r"^\d+$", lines[0]):
                continue
            timing = lines[1]
            if "-->" not in timing:
                continue
            text = "\n".join(lines[2:]).strip()
            if not text:
                continue
            start, end = [part.strip() for part in timing.split("-->", 1)]
            segments.append(
                {
                    "id": f"c{len(segments)}",
                    "index": len(segments),
                    "start": start,
                    "end": end,
                    "text": text,
                }
            )

    pages = _chunk_segments(segments)
    full = "\n".join(item["text"] for item in segments)
    meta = {"segments": segments, "format": suffix.lstrip(".")}
    return pages, enc, meta


def extract_spreadsheet(path: Path) -> tuple[list[str], str, dict]:
    suffix = path.suffix.lower()
    segments: list[dict] = []

    if suffix == ".csv":
        import csv

        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.reader(handle)
            for r_idx, row in enumerate(reader):
                for c_idx, value in enumerate(row):
                    text = str(value or "").strip()
                    if not text:
                        continue
                    segments.append(
                        {
                            "id": f"r{r_idx}c{c_idx}",
                            "row": r_idx,
                            "col": c_idx,
                            "text": text,
                        }
                    )
        pages = _chunk_segments(segments, 60)
        meta = {"segments": segments, "format": "csv"}
        return pages, "csv", meta

    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                text = str(cell.value or "").strip()
                if not text:
                    continue
                segments.append(
                    {
                        "id": f"s{sheet.title}!{cell.coordinate}",
                        "sheet": sheet.title,
                        "cell": cell.coordinate,
                        "text": text,
                    }
                )
    wb.close()
    pages = _chunk_segments(segments, 60)
    meta = {"segments": segments, "format": "xlsx"}
    return pages, "xlsx", meta


def extract_json_strings(path: Path) -> tuple[list[str], str, dict]:
    raw, enc = decode_text_bytes(path.read_bytes())
    data = json.loads(raw)
    segments: list[dict] = []

    def walk(value, path_key: str) -> None:
        if isinstance(value, str):
            text = value.strip()
            if text:
                segments.append({"id": path_key, "path": path_key, "text": text})
            return
        if isinstance(value, list):
            for idx, item in enumerate(value):
                walk(item, f"{path_key}[{idx}]")
            return
        if isinstance(value, dict):
            for key, item in value.items():
                child = f"{path_key}.{key}" if path_key else str(key)
                walk(item, child)

    walk(data, "")
    pages = _chunk_segments(segments, 50)
    meta = {"segments": segments, "format": "json", "root": data}
    return pages, enc, meta


def extract_html(path: Path) -> tuple[list[str], str, dict]:
    raw, enc = decode_text_bytes(path.read_bytes())
    text = re.sub(r"(?is)<(script|style).*?>.*?</\1>", " ", raw)
    text = re.sub(r"(?is)<br\s*/?>", "\n", text)
    text = re.sub(r"(?is)</p\s*>", "\n\n", text)
    text = re.sub(r"<[^>]+>", " ", text)
    text = html.unescape(re.sub(r"[ \\t]+", " ", text))
    text = re.sub(r"\n{3,}", "\n\n", text).strip()
    pages = [text] if text else [""]
    meta = {"segments": [{"id": "body", "text": text}], "format": "html", "raw_html": raw}
    return pages, enc, meta


def extract_epub(path: Path) -> tuple[list[str], str, dict]:
    try:
        from ebooklib import epub
    except ImportError as exc:
        raise RuntimeError("Install ebooklib: pip install ebooklib") from exc

    book = epub.read_epub(str(path))
    segments: list[dict] = []
    for idx, item in enumerate(book.get_items()):
        if item.get_type() != 9:  # ITEM_DOCUMENT
            continue
        raw = item.get_content().decode("utf-8", errors="replace")
        text = re.sub(r"(?is)<(script|style).*?>.*?</\1>", " ", raw)
        text = re.sub(r"<[^>]+>", " ", text)
        text = html.unescape(re.sub(r"\s+", " ", text)).strip()
        if not text:
            continue
        segments.append({"id": f"ch{idx}", "index": idx, "text": text, "href": item.get_name()})
    pages = _chunk_segments(segments, 20)
    full = "\n\n".join(item["text"] for item in segments)
    meta = {"segments": segments, "format": "epub"}
    return pages, "epub", meta


def _pages_non_empty(pages: list[str]) -> bool:
    return any(p.strip() for p in pages)


def _cell_text(cell) -> str:
    return str(cell or "").replace("|", "\\|").replace("\n", " ").strip()


def _table_to_markdown(table) -> str:
    try:
        raw = table.to_markdown().strip()
        if raw and "<br>" not in raw:
            return raw
    except (AttributeError, RuntimeError, ValueError, TypeError):
        pass

    try:
        rows = table.extract()
    except (AttributeError, RuntimeError, ValueError, TypeError):
        return ""

    if not rows:
        return ""

    grid = [[_cell_text(c) for c in row] for row in rows]
    width = max(len(row) for row in grid)
    if width == 0:
        return ""

    lines: list[str] = []
    header = grid[0] + [""] * (width - len(grid[0]))
    lines.append("| " + " | ".join(header[:width]) + " |")
    lines.append("| " + " | ".join(["---"] * width) + " |")
    for row in grid[1:]:
        padded = row + [""] * (width - len(row))
        lines.append("| " + " | ".join(padded[:width]) + " |")
    return "\n".join(lines)


def _page_to_markdown(page) -> str:
    parts: list[str] = []

    try:
        finder = page.find_tables()
        for table in getattr(finder, "tables", []) or []:
            md = _table_to_markdown(table)
            if md:
                parts.append(md)
    except (AttributeError, RuntimeError, ValueError, TypeError):
        pass

    try:
        data = page.get_text("dict")
        body_lines: list[str] = []
        for block in data.get("blocks", []):
            if block.get("type") != 0:
                continue
            for line in block.get("lines", []):
                spans = [span.get("text", "") for span in line.get("spans", [])]
                text = "".join(spans).rstrip()
                if text:
                    body_lines.append(text)
        body = "\n".join(body_lines).strip()
    except (AttributeError, RuntimeError, ValueError, TypeError):
        body = (page.get_text() or "").strip()

    if parts and body:
        return "\n\n".join(parts) + "\n\n" + body
    if parts:
        return "\n\n".join(parts)
    return body


def extract_pdf_pymupdf_markdown(path: Path) -> tuple[list[str], str] | None:
    try:
        import fitz  # pymupdf
    except ImportError:
        return None

    pages: list[str] = []
    with fitz.open(path) as doc:
        for page in doc:
            pages.append(_page_to_markdown(page))
    if not pages:
        pages = [""]
    if not _pages_non_empty(pages):
        return None
    return pages, "pymupdf-md"


def extract_pdf_pymupdf(path: Path) -> tuple[list[str], str] | None:
    try:
        import fitz  # pymupdf
    except ImportError:
        return None

    pages: list[str] = []
    with fitz.open(path) as doc:
        for page in doc:
            pages.append((page.get_text() or "").strip())
    if not pages:
        pages = [""]
    if not _pages_non_empty(pages):
        return None
    return pages, "pymupdf"


def extract_pdf_pypdf(path: Path) -> tuple[list[str], str] | None:
    try:
        from pypdf import PdfReader
    except ImportError:
        try:
            from PyPDF2 import PdfReader
        except ImportError:
            return None

    reader = PdfReader(str(path))
    pages = [(p.extract_text() or "").strip() for p in reader.pages]
    if not pages:
        pages = [""]
    if not _pages_non_empty(pages):
        return None
    return pages, "pypdf"


def _rects_overlap(a: tuple, b: tuple, tol: float = 3.0) -> bool:
    return not (a[2] < b[0] - tol or a[0] > b[2] + tol or a[3] < b[1] - tol or a[1] > b[3] + tol)


def extract_page_blocks(page) -> list:
    blocks: list = []
    table_rects: list[tuple] = []

    try:
        finder = page.find_tables()
        for table in list(getattr(finder, "tables", []) or [])[:4]:
            rows = [[str(cell or "").strip() for cell in row] for row in (table.extract() or [])]
            rows = [row for row in rows if any(cell for cell in row)]
            if rows:
                blocks.append({"kind": "table", "rows": rows})
                table_rects.append(tuple(table.bbox))
    except (AttributeError, RuntimeError, ValueError, TypeError):
        pass

    try:
        data = page.get_text("dict")
    except (AttributeError, RuntimeError, ValueError, TypeError):
        data = {"blocks": []}

    for block in data.get("blocks", []):
        if block.get("type") != 0:
            continue
        bbox = tuple(block.get("bbox", (0, 0, 0, 0)))
        if any(_rects_overlap(bbox, rect) for rect in table_rects):
            continue
        for line in block.get("lines", []):
            spans = line.get("spans", [])
            if not spans:
                continue
            text = "".join(span.get("text", "") for span in spans).strip()
            if not text:
                continue
            sizes = [float(span.get("size", 10.0)) for span in spans]
            max_size = max(sizes) if sizes else 10.0
            x0 = min(span.get("bbox", bbox)[0] for span in spans)

            if text.startswith(("•", "▪", "·", "-", "–")):
                blocks.append({"kind": "paragraph", "indent": max(0, int(x0 / 6)), "text": text, "level": 0})
                continue

            if max_size >= 18:
                blocks.append({"kind": "heading", "level": 1, "text": text})
            elif max_size >= 14:
                blocks.append({"kind": "heading", "level": 2, "text": text})
            elif max_size >= 12:
                blocks.append({"kind": "heading", "level": 3, "text": text})
            else:
                blocks.append(
                    {
                        "kind": "paragraph",
                        "indent": max(0, int((x0 - 48) / 5)),
                        "text": text,
                        "level": 0,
                    }
                )

    return blocks


def extract_pdf_structured(path: Path) -> tuple[list[str], list, str] | None:
    try:
        import fitz  # pymupdf
    except ImportError:
        return None

    pages: list[str] = []
    page_blocks: list = []
    with fitz.open(path) as doc:
        for page in doc:
            blocks = extract_page_blocks(page)
            page_blocks.append(blocks)
            plain_parts: list[str] = []
            for block in blocks:
                kind = block.get("kind")
                if kind == "table":
                    for row in block.get("rows", []):
                        plain_parts.append(" | ".join(row))
                else:
                    plain_parts.append(block.get("text", ""))
            pages.append("\n".join(p for p in plain_parts if p).strip())

    if not _pages_non_empty(pages):
        return None
    return pages, page_blocks, "structured"


def extract_pdf(path: Path, *, markdown: bool = False, structured: bool = False) -> tuple[list[str], str]:
    if structured:
        result = extract_pdf_structured(path)
        if result is not None:
            return result[0], result[2]
    if markdown:
        result = extract_pdf_pymupdf_markdown(path)
        if result is not None:
            return result

    for extractor in (extract_pdf_pymupdf, extract_pdf_pypdf):
        result = extractor(path)
        if result is not None:
            return result

    try:
        from PyPDF2 import PdfReader
    except ImportError as exc:
        raise RuntimeError("No PDF library available (install pymupdf or PyPDF2)") from exc

    reader = PdfReader(str(path))
    pages = [(p.extract_text() or "").strip() for p in reader.pages]
    if not pages:
        pages = [""]
    return pages, "pdf"


def emit(payload: dict) -> None:
    sys.stdout.buffer.write(json.dumps(payload, ensure_ascii=True).encode("utf-8"))
    sys.stdout.buffer.write(b"\n")
    sys.stdout.buffer.flush()


def emit_success(pages: list[str], encoding: str, workflow_id: str, workflow_meta: dict | None = None) -> int:
    full_text = "\n\n".join(p for p in pages if p.strip())
    if not full_text.strip():
        emit({"ok": False, "error": "No translatable text found in document"})
        return 2
    payload = {
        "ok": True,
        "text": full_text,
        "pages": pages,
        "page_count": len(pages),
        "encoding": encoding,
        "workflow_id": workflow_id,
    }
    if workflow_meta:
        payload["workflow_meta"] = workflow_meta
    emit(payload)
    return 0


def main() -> int:
    args = [a for a in sys.argv[1:] if a]
    markdown = False
    structured = False
    while args and args[0].startswith("--"):
        flag = args.pop(0)
        if flag == "--markdown":
            markdown = True
        elif flag == "--structured":
            structured = True

    if len(args) < 1:
        emit({"ok": False, "error": "usage: extract_document.py [--structured|--markdown] <file>"})
        return 2

    path = Path(args[0])
    if not path.is_file():
        emit({"ok": False, "error": f"file not found: {path}"})
        return 2

    suffix = path.suffix.lower()
    try:
        if suffix in {".txt", ".md"}:
            pages, encoding = extract_txt(path)
            workflow_id = "markdown" if suffix == ".md" else "text"
            return emit_success(pages, encoding, workflow_id)
        if suffix == ".docx":
            pages, encoding, meta = extract_docx(path)
            return emit_success(pages, encoding, "docx", meta)
        if suffix in {".srt", ".ass", ".vtt"}:
            pages, encoding, meta = extract_subtitles(path)
            return emit_success(pages, encoding, "subtitle", meta)
        if suffix in {".xlsx", ".csv"}:
            pages, encoding, meta = extract_spreadsheet(path)
            return emit_success(pages, encoding, "spreadsheet", meta)
        if suffix == ".json":
            pages, encoding, meta = extract_json_strings(path)
            return emit_success(pages, encoding, "json", meta)
        if suffix in {".html", ".htm"}:
            pages, encoding, meta = extract_html(path)
            return emit_success(pages, encoding, "html", meta)
        if suffix == ".epub":
            pages, encoding, meta = extract_epub(path)
            return emit_success(pages, encoding, "epub", meta)
        if suffix == ".pdf":
            if structured:
                packed = extract_pdf_structured(path)
                if packed is None:
                    emit({"ok": False, "error": "Structured PDF extraction failed"})
                    return 2
                pages, page_blocks, encoding = packed
                full_text = "\n\n".join(p for p in pages if p.strip())
                emit({
                    "ok": True,
                    "text": full_text,
                    "pages": pages,
                    "page_blocks": page_blocks,
                    "page_count": len(pages),
                    "encoding": encoding,
                    "workflow_id": "pdf_layout",
                })
                return 0
            pages, encoding = extract_pdf(path, markdown=markdown)
            full_text = "\n\n".join(p for p in pages if p.strip())
            if not full_text.strip():
                emit({
                    "ok": False,
                    "error": "No text layer in PDF (scan/image only). OCR is not supported yet.",
                })
                return 2
            emit({
                "ok": True,
                "text": full_text,
                "pages": pages,
                "page_count": len(pages),
                "encoding": encoding,
                "workflow_id": "pdf_layout",
            })
            return 0

        emit({"ok": False, "error": f"unsupported type: {suffix}"})
        return 2
    except Exception as exc:  # noqa: BLE001
        emit({"ok": False, "error": str(exc) or repr(exc)})
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
